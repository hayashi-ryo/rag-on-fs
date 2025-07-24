from openpyxl import load_workbook

def extract_xlsx_text(filepath: str) -> str:
    wb = load_workbook(filepath, data_only=True)
    lines = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows(values_only=True):
            row_text = [str(cell) if cell is not None else "" for cell in row]
            if any(row_text):
                lines.append("\t".join(row_text))
    return "\n".join(lines)
